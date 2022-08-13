from WindPy import w
import pandas as pd

import openpyxl

def get_evidence(sheet,evidence_start,evidence_end):
    evidence = []
    cell_tuples = sheet[evidence_start: evidence_end]
    for cells in cell_tuples:
        for cell in cells:
            evidence.append(cell.value)
    return evidence

def get_code(sheet,code_start,code_end):
    secucode = []
    cell_tuples = sheet[code_start: code_end]
    for cells in cell_tuples:
        for cell in cells:
            secucode.append(cell.value)
    return secucode

def savedata(data2021 ,save_address,evidence):
    workbook = openpyxl.Workbook()
    sheet0 = workbook.create_sheet(index=0)  # 创建sheet0
    #sheet5 = workbook.create_sheet(index=5)
    for e in range(len(evidence)):
        sheet0.cell(1 , e+1).value = evidence[e]
    m=len(data2021)
    n=len(data2021[0])
    for i in range(n):
        for j in range(m):
            sheet0.cell(i + 2, j + 1).value = data2021[j][i]
            #sheet5.cell(i + 1, j + 1).value = data2021[j][i]
    workbook.save(save_address)  # 保存文件


def getdata(code):
    #获得公式法指标
    evidence_start='F2'
    evidence_end='SN2'
    wb = openpyxl.load_workbook(r'公式法提取财报模板.xlsx')
    sheet = wb['Sheet1']
    evidence = get_evidence(sheet,evidence_start,evidence_end)
    print(evidence)
    print('请输入报告年份（直接输年份如2021）:')
    year=str(input())
    astring="unit=1;rptDate="+year+"1231;rptType=1;year="+year
    data = w.wss(code, evidence, astring)
    #save_address=r'C:\Users\markmxu\Desktop\公式法提取.xlsx'
    #savedata(data.Data,save_address)
    return [data.Data , evidence]
