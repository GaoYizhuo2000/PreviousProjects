from WindPy import w
import pandas as pd

import openpyxl

def get_evidence(sheet):
    evidence = []
    cell_tuples = sheet['PM2': 'QA2']
    for cells in cell_tuples:
        for cell in cells:
            evidence.append(cell.value)
    return evidence

def get_code(sheet,code_start,code_end):
    secucode = []
    cell_tuples = sheet[code_start: code_end]
    for cells in cell_tuples:
        for cell in cells:
            secucode.append(cell.value)
    return secucode

def savedata(data,save_address):
    workbook = openpyxl.Workbook()
    sheet0 = workbook.create_sheet(index=0)  # 创建sheet0
    #sheet5 = workbook.create_sheet(index=5)
    m=len(data)
    n=len(data[0])
    for i in range(n):
        for j in range(m):
            sheet0.cell(i + 1, j + 1).value = data[j][i]
            #sheet5.cell(i + 1, j + 1).value = data2021[j][i]
    workbook.save(save_address)  # 保存文件

def getdata(code , year):
    wb = openpyxl.load_workbook(r'十六期170家主体485个字段.xlsx')
    sheet = wb['2016']
    #获取主体公司债券
    evidence=get_evidence(sheet)
    print('导入指标如下所示：')
    print(evidence)
    astring = "unit=1;rptDate=" + year + "1231;rptType=2"
    data = w.wss(code, evidence,astring)
    return [data.Data , evidence]

