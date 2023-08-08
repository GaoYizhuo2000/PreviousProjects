from WindPy import w
import pandas as pd
import openpyxl
import  system415
import main80
import mainxx
import windannex
def get_code(sheet):
    secucode = []
    i = 0
    while 1:
        i +=1
        if sheet.cell(row = i , column = 1).value :
            secucode.append(sheet.cell(row = i , column = 1).value)
        else:
            break
    #cell_tuples = sheet[code_start: code_end]     之前的代码
    #for cells in cell_tuples:
        #for cell in cells:
            #secucode.append(cell.value)
    return secucode
#def sure_save(data , evidence):
   # print('是否进行xlsx文件保存？（Y/N）')
   # sure = input()
   # if sure == 'Y' or sure == 'y':
     #   print('输入保存路径(或直接输入文件名，加后缀.xlsx)：')
    #    save_address = str(input())
     #   system415.savedata(data, save_address , evidence)
       # print('已保存')
   # else:
    #    pass
def sure_exceed(data):
    if data[0][0] == 'CWSSService: quota exceeded.':
        print('Markmxu提醒您，提取配额已超量，请充值~')
        pass
    else:
        print(data)
        pass

if __name__ == '__main__':
    w.start()  # 默认命令超时时间为120秒，
    w.isconnected()  # 判断WindPy是否已经登录成功
    print('----------------------------------------------------------------欢迎使用wind财报智评提数系统1.0----------------------------------------------------------------------------------------------')
    print('----------------------------------------------------------------------Writen by neroGao--------------------------------------------------------------------------------------------------')
    while 1:
        xxxx = input('继续输入‘1’，结束输入‘0’：')
        if xxxx != '1':
            break
        print('把待提取的股票代码放到  ’__待提取股票代码__.xlsx‘ 的Sheet1里并保存，放在第一列')
        xx = input('保存好后输入任意值继续：')
        wb = openpyxl.load_workbook('__待提取股票代码__.xlsx')
        sheet = wb['Sheet1']
        # 获取主体公司债券
        code = get_code(sheet)
        print('已选择债券主体：')
        print(code)

        rptyear = input('输入报告年份（如2021）：')
        r1 = system415.getdata(code , rptyear)  #公式法提取1
        data1 = r1[0]                   #公式法提取2
        evidence1 = r1[1]
        r2 = main80.getdata(code , rptyear) #公式法提取2
        data2 = r2[0]
        evidence2 = r2[1]
        r3 = mainxx.getdata(code , rptyear)  #母公司提取
        data3 = r3[0]
        evidence3 = r3[1]
        r4 = windannex.getdata(code , rptyear)  #附加提取
        data4 = r4[0]
        evidence4 = r4[1]
        evidencelist = evidence1 + evidence2 + evidence3 + evidence4

        print('输入保存路径(或直接输入文件名，加后缀.xlsx)：')
        save_address = str(input())
        system415.savedata1(r1[0], save_address , evidencelist)
        system415.savedata2(r2[0], save_address)     #改函数把2后的都改成读取1的表并写入
        system415.savedata3(r3[0], save_address)
        system415.savedata4(r4[0], save_address)




        
