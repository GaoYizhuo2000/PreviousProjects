from WindPy import w
import pandas as pd
import openpyxl
import  system415
import main80
import mainxx
import windannex
def get_code(sheet):
    secucode = []
    i = 0
    while 1:
        i +=1
        if sheet.cell(row = i , column = 1).value :
            secucode.append(sheet.cell(row = i , column = 1).value)
        else:
            break
    #cell_tuples = sheet[code_start: code_end]     之前的代码
    #for cells in cell_tuples:
        #for cell in cells:
            #secucode.append(cell.value)
    return secucode
def sure_save(data , evidence):
    print('是否进行xlsx文件保存？（Y/N）')
    sure = input()
    if sure == 'Y' or sure == 'y':
        print('输入保存路径：')
        save_address = str(input())
        system415.savedata(data, save_address , evidence)
        print('已保存')
    else:
        pass
def sure_exceed(data):
    if data[0][0] == 'CWSSService: quota exceeded.':
        print('Markmxu提醒您，提取配额已超量，请充值~')
        pass
    else:
        print(data)
        pass

if __name__ == '__main__':
    w.start()  # 默认命令超时时间为120秒，
    w.isconnected()  # 判断WindPy是否已经登录成功
    print('----------------------------------------------------------------欢迎使用wind财报智评提数系统1.0----------------------------------------------------------------------------------------------')
    print('----------------------------------------------------------------------Writen by neroGao--------------------------------------------------------------------------------------------------')
    print('把待提取的股票代码放到  ’__待提取股票代码__.xlsx‘ 的Sheet1里并保存，放在第一列')
    xx = input('保存好后输入任意值继续：')
    wb = openpyxl.load_workbook('__待提取股票代码__.xlsx')
    sheet = wb['Sheet1']
    # 获取主体公司债券
    code = get_code(sheet)
    print('已选择债券主体：')
    print(code)
    #选择提数功能
    print('提数功能：1--公式法   2--80公式   3--母公司公司   4--附加wind   其他键--退出')
    while 1:
        print('请输入提数功能：')
        button = str(input())
        if button == '1':
            r1 = system415.getdata(code)
            data=r1[0]
            evidence = r1[1]
            print('你选择了公式法提取')
            sure_exceed(data)
            sure_save(data ,evidence)
        elif button == '2':
            r2 =  main80.getdata(code)
            data = r2[0]
            evidence = r2[1]
            print('你选择了80公式提取')
            sure_exceed(data)
            sure_save(data , evidence)
        elif button == '3':
            r3 = mainxx.getdata(code)
            data = r3[0]
            evidence = r3[1]
            print('你选择了母公司报表提取')
            sure_exceed(data)
            sure_save(data , evidence)
        elif button == '4':
            r4 = windannex.getdata(code)
            data = r4[0]
            evidence = r4[1]
            print('你选择了附加万德提取')
            sure_exceed(data)
            sure_save(data , evidence)
        else:
            print('退出！')
            break


        
