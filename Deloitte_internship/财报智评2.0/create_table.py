from WindPy import w
import pandas as pd
from matplotlib import  pyplot as plt
import openpyxl

def create(save_address,sheet1,sheet2):
    evidence=[]
    code=[]
    company=[]
    workbook = openpyxl.Workbook()
    sheet0 = workbook.create_sheet(index=0)  # 创建sheet0

    cell_tuples = sheet1['A1': 'ZU1']
    for cells in cell_tuples:
        for cell in cells:
            evidence.append(cell.value)
    m=len(evidence)
    for i in range(m):
        sheet0.cell(1, i + 1).value = evidence[i]
    #导入债券
    cellcode=sheet2['A2': 'A13']
    for cells in cellcode:
        for cell in cells:
            code.append(cell.value)
    m=len(code)
    for i in range(m):
        sheet0.cell(i+2, 4).value = code[i]
    #导入公司
    cellcompany = sheet2['C2': 'C13']
    for cells in cellcompany:
        for cell in cells:
            company.append(cell.value)
    m = len(company)
    for i in range(m):
        sheet0.cell(i + 2, 1).value = company[i]

    workbook.save(save_address)  # 保存文件
if __name__ == '__main__':
    wb = openpyxl.load_workbook(r'C:\Users\markmxu\Desktop\20220607\财报智评0607.xlsx')
    sheet1 = wb['Sheet1']
    ws = openpyxl.load_workbook(r'C:\Users\markmxu\Desktop\20220607\财报智评新上市新发债：全量A股＋全量信用债0606.xlsx')
    sheet2 = ws['工作表1']
    save_address = r'C:\Users\markmxu\Desktop\财报智评06.xlsx'
    create(save_address,sheet1,sheet2)
