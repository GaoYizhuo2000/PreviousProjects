import openpyxl

#def get_code(sheet):
    #secucode = []
    #i = 0
    #while 1:
        # +=1
        #if sheet.cell(row = i , column = 1).value :
            #secucode.append(sheet.cell(row = i , column = 1).value)
        #else:
            #break
    #return secucode

#print('把待提取的股票代码放到  ’__待提取股票代码__.xlsx‘ 的Sheet1里并保存，放在第一列')
#xx = input('保存好后输入任意值继续：')
#wb = openpyxl.load_workbook('__待提取股票代码__.xlsx')
#sheet = wb['Sheet1']
# 获取主体公司债券
#code = get_code(sheet)
#print('已选择债券主体：')
#print(code)
def a():
    list1 = [1,2,3]
    list2 = [4,5,6]
    return [list1 , list2]
b = a()
print(b)
print(b[0])
