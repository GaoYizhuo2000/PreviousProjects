from WindPy import w
import pandas as pd

import openpyxl
from eighty.fomula80_2016 import wss_get2016
from eighty.fomula80_2017 import wss_get2017
from eighty.fomula80_2018 import wss_get2018
from eighty.fomula80_2019 import wss_get2019
from eighty.fomula80_2020 import wss_get2020
from eighty.fomula80_2021 import wss_get2021


def get_evidence(sheet):
    evidence = []
    cell_tuples = sheet['SO2': 'VP2']
    for cells in cell_tuples:
        for cell in cells:
            evidence.append(cell.value)
    return evidence

def get_code(sheet,code_start,code_end):
    secucode = []
    cell_tuples = sheet[code_start: code_end]
    for cells in cell_tuples:
        for cell in cells:
            secucode.append(cell.value)
    return secucode

def savedata(data,save_address):
    workbook = openpyxl.Workbook()
    sheet0 = workbook.create_sheet(index=0)  # 创建sheet0
    # sheet5 = workbook.create_sheet(index=5)
    m = len(data)
    n = len(data[0])
    for i in range(n):
        for j in range(m):
            sheet0.cell(i + 1, j + 1).value = data[j][i]
            # sheet5.cell(i + 1, j + 1).value = data2021[j][i]
    workbook.save(save_address)  # 保存文件


def wss_getYear(code,evidence,year):
    if year==2016:
        data=wss_get2016(code)
    elif year==2017:
        data = wss_get2017(code)
    elif year==2018:
        data = wss_get2018(code)
    elif year==2019:
        data = wss_get2019(code)
    elif year==2020:
        data = wss_get2020(code)
    else:
        data = wss_get2021(code)
    return data

def trasform(data):
    DATA=[]
    m=len(data)
    for i in range(m):
        n=len(data[i])
        for j in range(n):
            DATA.append(data[i][j])
    return DATA


def getdata(code):
    wb = openpyxl.load_workbook(r'公式法提取财报模板.xlsx')
    sheet = wb['Sheet1']
    # 获取evidence
    evidence = get_evidence(sheet)
    print('导入指标如下所示：')
    print(evidence)
    print('请输入报告年份（直接输年份如2021）:')
    year = input()
    data=wss_getYear(code,evidence,year) #80字段
    data=trasform(data)
    return [data , evidence]
